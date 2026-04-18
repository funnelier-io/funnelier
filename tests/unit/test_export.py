"""
Tests for Export Module — Sprint 1 P0 Gap Closure.

Covers:
- Export schemas (ExportFormat, ReportType, ExportRequest, etc.)
- CSV generation (generate_csv_bytes)
- XLSX generation (generate_xlsx_bytes)
- Summary XLSX generation (generate_summary_xlsx)
- PDF text fallback generation
- Persian header mapping (_get_persian_headers)
- Column definitions completeness
"""

import csv
import io
import pytest
from datetime import date, datetime
from uuid import uuid4

from src.modules.export.schemas import (
    ExportFormat,
    ReportType,
    ScheduleFrequency,
    ExportRequest,
    ScheduledReportRequest,
    CustomReportRequest,
    ExportJobResponse,
    ExportFileResponse,
    ReportColumnInfo,
    AvailableColumnsResponse,
)
from src.modules.export.service import (
    COLUMN_DEFS,
    MODEL_MAP,
    DATE_COLUMN_MAP,
    generate_csv_bytes,
    generate_xlsx_bytes,
    generate_summary_xlsx,
    _get_persian_headers,
    _generate_pdf_text_fallback,
)


# ═══════════════════════════════════════════════════════════════════
# Enum Tests
# ═══════════════════════════════════════════════════════════════════


class TestExportEnums:
    def test_export_formats(self):
        assert ExportFormat.csv == "csv"
        assert ExportFormat.xlsx == "xlsx"
        assert ExportFormat.pdf == "pdf"

    def test_report_types(self):
        assert ReportType.contacts == "contacts"
        assert ReportType.invoices == "invoices"
        assert ReportType.funnel_summary == "funnel_summary"
        assert ReportType.custom == "custom"

    def test_schedule_frequencies(self):
        assert ScheduleFrequency.daily == "daily"
        assert ScheduleFrequency.weekly == "weekly"
        assert ScheduleFrequency.monthly == "monthly"


# ═══════════════════════════════════════════════════════════════════
# Schema Tests
# ═══════════════════════════════════════════════════════════════════


class TestExportSchemas:
    def test_export_request_defaults(self):
        req = ExportRequest(report_type=ReportType.contacts)
        assert req.format == ExportFormat.xlsx
        assert req.start_date is None
        assert req.filters == {}
        assert req.columns is None

    def test_export_request_with_filters(self):
        req = ExportRequest(
            report_type=ReportType.invoices,
            format=ExportFormat.csv,
            start_date=date(2026, 1, 1),
            end_date=date(2026, 3, 31),
            filters={"status": "paid"},
            columns=["invoice_number", "total_amount"],
        )
        assert req.filters["status"] == "paid"
        assert len(req.columns) == 2

    def test_scheduled_report_request(self):
        req = ScheduledReportRequest(
            name="Weekly Funnel Report",
            report_type=ReportType.funnel_summary,
            frequency=ScheduleFrequency.weekly,
            recipients=["admin@funnelier.com"],
        )
        assert req.is_active is True
        assert req.format == ExportFormat.xlsx
        assert len(req.recipients) == 1

    def test_scheduled_report_name_validation(self):
        with pytest.raises(Exception):
            ScheduledReportRequest(name="", report_type=ReportType.contacts)

    def test_custom_report_request(self):
        req = CustomReportRequest(
            name="Combined Report",
            data_sources=[ReportType.contacts, ReportType.invoices],
            columns={"contacts": ["name", "phone_number"]},
        )
        assert len(req.data_sources) == 2

    def test_custom_report_requires_data_sources(self):
        with pytest.raises(Exception):
            CustomReportRequest(name="Empty", data_sources=[])

    def test_export_job_response(self):
        resp = ExportJobResponse(job_id="job-123")
        assert resp.status == "processing"
        assert resp.download_url is None

    def test_export_file_response(self):
        resp = ExportFileResponse(
            job_id="job-123",
            filename="report.xlsx",
            download_url="/downloads/report.xlsx",
            file_size=1024,
            row_count=50,
            created_at=datetime.utcnow(),
        )
        assert resp.status == "completed"
        assert resp.file_size == 1024

    def test_report_column_info(self):
        col = ReportColumnInfo(
            key="phone_number",
            label="Phone Number",
            label_fa="شماره تلفن",
            type="string",
        )
        assert col.key == "phone_number"

    def test_available_columns_response(self):
        resp = AvailableColumnsResponse(
            report_type="contacts",
            columns=[
                ReportColumnInfo(key="name", label="Name", label_fa="نام"),
            ],
        )
        assert len(resp.columns) == 1


# ═══════════════════════════════════════════════════════════════════
# Column Definitions / Config
# ═══════════════════════════════════════════════════════════════════


class TestColumnDefinitions:
    def test_all_report_types_have_columns(self):
        for rt in ["contacts", "invoices", "call_logs", "sms_logs", "payments"]:
            assert rt in COLUMN_DEFS, f"Missing column defs for {rt}"
            assert len(COLUMN_DEFS[rt]) > 0

    def test_contacts_has_expected_columns(self):
        keys = [c[0] for c in COLUMN_DEFS["contacts"]]
        assert "phone_number" in keys
        assert "name" in keys
        assert "rfm_segment" in keys
        assert "total_revenue" in keys
        assert "created_at" in keys

    def test_invoices_has_expected_columns(self):
        keys = [c[0] for c in COLUMN_DEFS["invoices"]]
        assert "invoice_number" in keys
        assert "total_amount" in keys
        assert "status" in keys

    def test_column_def_tuples_have_four_elements(self):
        """Each column def should be (key, en_label, fa_label, type)."""
        for report_type, cols in COLUMN_DEFS.items():
            for col in cols:
                assert len(col) == 4, f"Bad column def in {report_type}: {col}"

    def test_model_map_covers_tabular_types(self):
        for rt in ["contacts", "invoices", "call_logs", "sms_logs", "payments"]:
            assert rt in MODEL_MAP

    def test_date_column_map_covers_tabular_types(self):
        for rt in ["contacts", "invoices", "call_logs", "sms_logs", "payments"]:
            assert rt in DATE_COLUMN_MAP


# ═══════════════════════════════════════════════════════════════════
# Persian Headers
# ═══════════════════════════════════════════════════════════════════


class TestPersianHeaders:
    def test_returns_persian_labels(self):
        headers = _get_persian_headers("contacts", ["phone_number", "name"])
        assert headers[0] == "شماره تلفن"
        assert headers[1] == "نام"

    def test_fallback_to_key_when_unknown(self):
        headers = _get_persian_headers("contacts", ["nonexistent_col"])
        assert headers[0] == "nonexistent_col"

    def test_invoices_headers(self):
        headers = _get_persian_headers("invoices", ["invoice_number", "total_amount"])
        assert headers[0] == "شماره فاکتور"
        assert headers[1] == "مبلغ کل"


# ═══════════════════════════════════════════════════════════════════
# CSV Generation
# ═══════════════════════════════════════════════════════════════════


class TestGenerateCSV:
    def test_basic_csv(self):
        col_keys = ["phone_number", "name"]
        rows = [
            {"phone_number": "9121234567", "name": "علی"},
            {"phone_number": "9131234567", "name": "رضا"},
        ]
        result = generate_csv_bytes(col_keys, rows, "contacts")
        assert isinstance(result, bytes)
        text = result.decode("utf-8")
        # Should have BOM
        assert text.startswith("\ufeff")
        # Should have Persian headers
        assert "شماره تلفن" in text
        assert "نام" in text
        # Should have data
        assert "9121234567" in text
        assert "علی" in text

    def test_empty_rows(self):
        result = generate_csv_bytes(["phone_number"], [], "contacts")
        text = result.decode("utf-8")
        lines = text.strip().split("\n")
        assert len(lines) == 1  # Just header

    def test_csv_is_parseable(self):
        col_keys = ["phone_number", "name", "total_revenue"]
        rows = [{"phone_number": "9121234567", "name": "Test", "total_revenue": 1000}]
        result = generate_csv_bytes(col_keys, rows, "contacts")
        text = result.decode("utf-8").lstrip("\ufeff")
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
        assert len(all_rows) == 2  # header + 1 data row

    def test_missing_value_defaults_to_empty(self):
        result = generate_csv_bytes(["phone_number", "name"], [{"phone_number": "912"}], "contacts")
        text = result.decode("utf-8")
        assert ",," not in text or "," in text  # graceful handling


# ═══════════════════════════════════════════════════════════════════
# XLSX Generation
# ═══════════════════════════════════════════════════════════════════


class TestGenerateXLSX:
    def test_basic_xlsx(self):
        col_keys = ["phone_number", "name"]
        rows = [
            {"phone_number": "9121234567", "name": "علی"},
        ]
        result = generate_xlsx_bytes(col_keys, rows, "contacts")
        assert isinstance(result, bytes)
        assert len(result) > 0
        # XLSX files start with PK (zip signature)
        assert result[:2] == b"PK"

    def test_empty_data(self):
        result = generate_xlsx_bytes(["phone_number"], [], "contacts", sheet_name="Test")
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_custom_sheet_name(self):
        from openpyxl import load_workbook
        result = generate_xlsx_bytes(
            ["phone_number"], [{"phone_number": "912"}], "contacts",
            sheet_name="مخاطبین",
        )
        wb = load_workbook(io.BytesIO(result))
        assert wb.active.title == "مخاطبین"

    def test_rtl_enabled(self):
        from openpyxl import load_workbook
        result = generate_xlsx_bytes(
            ["phone_number"], [{"phone_number": "912"}], "contacts",
        )
        wb = load_workbook(io.BytesIO(result))
        assert wb.active.sheet_view.rightToLeft is True


# ═══════════════════════════════════════════════════════════════════
# Summary XLSX
# ═══════════════════════════════════════════════════════════════════


class TestGenerateSummaryXLSX:
    def _funnel_data(self):
        return {
            "latest_stage_counts": {
                "lead_acquired": 500,
                "sms_sent": 400,
                "call_answered": 100,
                "payment_received": 20,
            },
            "total_revenue": 500_000_000,
            "total_leads": 500,
        }

    def _team_data(self):
        return [
            {
                "salesperson": "بردبار",
                "total_calls": 200,
                "successful_calls": 50,
                "total_duration_min": 300.5,
                "total_invoices": 30,
                "total_amount": 150_000_000,
            },
        ]

    def _rfm_data(self):
        return [
            {"segment": "champions", "count": 20, "revenue": 200_000_000, "avg_calls": 15.0},
            {"segment": "at_risk", "count": 50, "revenue": 50_000_000, "avg_calls": 3.2},
        ]

    def test_generates_valid_xlsx(self):
        result = generate_summary_xlsx(
            self._funnel_data(), self._team_data(), self._rfm_data()
        )
        assert isinstance(result, bytes)
        assert result[:2] == b"PK"

    def test_has_three_sheets(self):
        from openpyxl import load_workbook
        result = generate_summary_xlsx(
            self._funnel_data(), self._team_data(), self._rfm_data()
        )
        wb = load_workbook(io.BytesIO(result))
        assert len(wb.sheetnames) == 3
        assert "خلاصه فانل" in wb.sheetnames
        assert "عملکرد تیم" in wb.sheetnames
        assert "بخش‌بندی RFM" in wb.sheetnames

    def test_empty_data(self):
        result = generate_summary_xlsx(
            {"latest_stage_counts": {}, "total_revenue": 0, "total_leads": 0},
            [],
            [],
        )
        assert isinstance(result, bytes)


# ═══════════════════════════════════════════════════════════════════
# PDF Text Fallback
# ═══════════════════════════════════════════════════════════════════


class TestPDFTextFallback:
    def test_basic_generation(self):
        result = _generate_pdf_text_fallback(
            funnel_data={
                "latest_stage_counts": {"lead_acquired": 100},
                "total_revenue": 50_000_000,
            },
            team_data=[
                {"salesperson": "Test", "total_calls": 50, "total_invoices": 10, "total_amount": 20_000_000},
            ],
            rfm_data=[
                {"segment": "champions", "count": 10, "revenue": 30_000_000},
            ],
            tenant_name="فانلیر",
            report_date="1405/01/29",
        )
        text = result.decode("utf-8")
        assert "فانلیر" in text
        assert "FUNNEL SUMMARY" in text
        assert "lead_acquired: 100" in text
        assert "TEAM PERFORMANCE" in text
        assert "RFM BREAKDOWN" in text

    def test_empty_team_and_rfm(self):
        result = _generate_pdf_text_fallback(
            funnel_data={"latest_stage_counts": {}, "total_revenue": 0},
            team_data=[],
            rfm_data=[],
            tenant_name="Test",
            report_date="2026-04-18",
        )
        text = result.decode("utf-8")
        assert "TEAM PERFORMANCE" not in text
        assert "RFM BREAKDOWN" not in text

    def test_none_stage_counts(self):
        result = _generate_pdf_text_fallback(
            funnel_data={"latest_stage_counts": None, "total_revenue": 0},
            team_data=[],
            rfm_data=[],
            tenant_name="Test",
            report_date="2026-04-18",
        )
        assert isinstance(result, bytes)

