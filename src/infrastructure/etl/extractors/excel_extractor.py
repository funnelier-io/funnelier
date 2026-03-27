"""
Excel Extractor

Extracts data from Excel files (.xlsx, .xls), supporting:
- Multiple sheets
- Lead data with category information
- Batch processing
"""

from datetime import datetime
from pathlib import Path
from typing import Any, AsyncIterator

import openpyxl

from src.core.utils import normalize_phone_strict
from openpyxl.utils import get_column_letter

from src.core.interfaces import DataRecord, FileSourceConfig

from .base import BaseExtractor, ExtractorRegistry


@ExtractorRegistry.register("excel")
class ExcelExtractor(BaseExtractor):
    """
    Extractor for Excel files.
    Primary use case: Lead lists categorized by source/campaign.
    """

    def __init__(self, config: FileSourceConfig, tenant_id: str | None = None):
        super().__init__(config, tenant_id)
        self._file_config: FileSourceConfig = config
        self._files: list[Path] = []
        self._workbooks: dict[Path, openpyxl.Workbook] = {}

    @property
    def source_type(self) -> str:
        return "excel"

    async def connect(self) -> bool:
        """Discover Excel files matching pattern."""
        try:
            if self._file_config.file_path:
                path = Path(self._file_config.file_path)
                if path.exists() and path.suffix in (".xlsx", ".xls"):
                    self._files = [path]
                else:
                    return False
            elif self._file_config.file_pattern:
                base_path = Path(self._file_config.file_pattern).parent
                pattern = Path(self._file_config.file_pattern).name
                if base_path.exists():
                    self._files = [
                        f
                        for f in base_path.glob(pattern)
                        if f.suffix in (".xlsx", ".xls")
                    ]
                else:
                    self._files = [
                        f
                        for f in Path.cwd().glob(self._file_config.file_pattern)
                        if f.suffix in (".xlsx", ".xls")
                    ]

            self._connected = bool(self._files)
            return self._connected
        except Exception:
            return False

    async def disconnect(self) -> None:
        """Close workbooks and clean up."""
        for wb in self._workbooks.values():
            wb.close()
        self._workbooks.clear()
        self._files = []
        self._connected = False

    async def test_connection(self) -> tuple[bool, str]:
        """Test if Excel files are accessible."""
        if not self._files:
            success = await self.connect()
            if not success:
                return False, "No Excel files found matching pattern"

        for file_path in self._files:
            if not file_path.exists():
                return False, f"File not found: {file_path}"
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                wb.close()
            except Exception as e:
                return False, f"Cannot read {file_path}: {str(e)}"

        return True, f"Found {len(self._files)} Excel file(s)"

    async def extract(
        self,
        batch_size: int = 1000,
        **kwargs: Any,
    ) -> AsyncIterator[list[DataRecord]]:
        """Extract records from Excel files in batches."""
        if not self._connected:
            await self.connect()

        sheet_name = kwargs.get("sheet_name")

        for file_path in self._files:
            async for batch in self._extract_file(file_path, batch_size, sheet_name):
                yield batch

    async def _extract_file(
        self,
        file_path: Path,
        batch_size: int,
        sheet_name: str | None = None,
    ) -> AsyncIterator[list[DataRecord]]:
        """Extract records from a single Excel file."""
        # Load workbook (data_only=True to get calculated values)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        self._workbooks[file_path] = wb

        # Get sheets to process
        sheets = [sheet_name] if sheet_name else wb.sheetnames

        for sheet in sheets:
            if sheet not in wb.sheetnames:
                continue

            ws = wb[sheet]
            batch: list[DataRecord] = []

            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append(f"column_{cell.column}")

            # Extract data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):  # Skip empty rows
                    continue

                row_data = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        field_name = headers[col_idx]
                        row_data[field_name] = self._clean_value(value)

                # Add metadata
                row_data["_source_file"] = file_path.name
                row_data["_source_sheet"] = sheet
                row_data["_category"] = self._extract_category(file_path)

                record = self._create_record(
                    data=row_data,
                    raw_data={
                        "file": str(file_path),
                        "sheet": sheet,
                        "row": row_idx,
                    },
                )
                batch.append(record)

                if len(batch) >= batch_size:
                    yield batch
                    batch = []

            if batch:
                yield batch

    def _clean_value(self, value: Any) -> Any:
        """Clean and normalize cell value."""
        if value is None:
            return None
        if isinstance(value, str):
            return value.strip() or None
        if isinstance(value, datetime):
            return value.isoformat()
        return value

    def _extract_category(self, file_path: Path) -> str:
        """Extract category from filename."""
        # Remove extension and clean up
        name = file_path.stem
        # Remove common prefixes
        for prefix in ["«", "»", "لید", "سرنخ"]:
            name = name.replace(prefix, "")
        return name.strip()

    async def get_schema(self) -> dict[str, Any]:
        """Get schema from first file."""
        if not self._files:
            await self.connect()

        if not self._files:
            return {}

        wb = openpyxl.load_workbook(self._files[0], read_only=True)

        schemas = {}
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            headers = [
                str(cell.value).strip() if cell.value else f"column_{i}"
                for i, cell in enumerate(ws[1], 1)
            ]
            schemas[sheet] = {"fields": headers}

        wb.close()

        return {
            "file_count": len(self._files),
            "sheets": schemas,
        }

    async def get_record_count(self) -> int | None:
        """Get approximate record count from all files."""
        if not self._files:
            await self.connect()

        total = 0
        for file_path in self._files:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                total += max(0, ws.max_row - 1)  # Subtract header row
            wb.close()

        return total


class LeadExcelExtractor(ExcelExtractor):
    """
    Specialized extractor for lead list Excel files.
    Extracts and normalizes lead data with category tagging.
    """

    async def _extract_file(
        self,
        file_path: Path,
        batch_size: int,
        sheet_name: str | None = None,
    ) -> AsyncIterator[list[DataRecord]]:
        """Extract and normalize lead records."""
        async for batch in super()._extract_file(file_path, batch_size, sheet_name):
            normalized_batch = []
            for record in batch:
                data = record.data
                normalized = {
                    "phone_number": self._normalize_phone(self._find_phone(data)),
                    "name": data.get("نام") or data.get("name") or data.get("نام و نام خانوادگی"),
                    "company": data.get("شرکت") or data.get("company"),
                    "city": data.get("شهر") or data.get("city"),
                    "category": data.get("_category"),
                    "source_file": data.get("_source_file"),
                    "source_sheet": data.get("_source_sheet"),
                    "raw_data": data,
                }
                normalized_batch.append(self._create_record(normalized, record.raw_data))
            yield normalized_batch

    def _find_phone(self, data: dict[str, Any]) -> str | None:
        """Find phone number from various possible field names."""
        phone_fields = [
            "شماره",
            "تلفن",
            "موبایل",
            "phone",
            "mobile",
            "number",
            "شماره تماس",
            "شماره موبایل",
            "شماره تلفن",
        ]
        for field in phone_fields:
            if field in data and data[field]:
                return str(data[field])
        # Check for fields containing phone keywords
        for key, value in data.items():
            if value and any(pf in key.lower() for pf in ["phone", "mobile", "شماره", "تلفن"]):
                return str(value)
        return None

    def _normalize_phone(self, phone: str | None) -> str | None:
        """Normalize phone number to standard format."""
        if not phone:
            return None
        return normalize_phone_strict(str(phone))

