"""
Export & Reporting — API Routes

Endpoints for CSV/XLSX/PDF export, scheduled reports, and custom report builder.
"""

from __future__ import annotations

import logging
from datetime import date, datetime
from typing import Annotated, Any
from uuid import UUID, uuid4

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from src.api.dependencies import get_current_tenant_id, get_db_session
from src.modules.auth.api.routes import require_auth

from src.modules.export.schemas import (
    AvailableColumnsResponse,
    CustomReportRequest,
    ExportFileResponse,
    ExportFormat,
    ExportJobResponse,
    ExportRequest,
    ReportColumnInfo,
    ReportType,
    ScheduledReportListResponse,
    ScheduledReportRequest,
    ScheduledReportResponse,
)
from src.modules.export.service import (
    COLUMN_DEFS,
    fetch_export_rows,
    fetch_funnel_summary,
    fetch_rfm_breakdown,
    fetch_team_performance,
    generate_csv_bytes,
    generate_pdf_bytes,
    generate_summary_xlsx,
    generate_xlsx_bytes,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/export", tags=["Export & Reporting"])


# ═════════════════════════════════════════════════════════════════════════
# Metadata: available columns per report type
# ═════════════════════════════════════════════════════════════════════════

@router.get("/columns/{report_type}", response_model=AvailableColumnsResponse)
async def get_available_columns(
    report_type: str,
):
    """Return available columns for a report type."""
    col_defs = COLUMN_DEFS.get(report_type)
    if col_defs is None:
        raise HTTPException(status_code=404, detail=f"Unknown report type: {report_type}")

    return AvailableColumnsResponse(
        report_type=report_type,
        columns=[
            ReportColumnInfo(key=k, label=en, label_fa=fa, type=tp)
            for k, en, fa, tp in col_defs
        ],
    )


@router.get("/columns", response_model=dict[str, list[ReportColumnInfo]])
async def get_all_columns():
    """Return all available columns grouped by report type."""
    result: dict[str, list[ReportColumnInfo]] = {}
    for rt, defs in COLUMN_DEFS.items():
        result[rt] = [
            ReportColumnInfo(key=k, label=en, label_fa=fa, type=tp)
            for k, en, fa, tp in defs
        ]
    return result


# ═════════════════════════════════════════════════════════════════════════
# Tabular export: contacts, invoices, call_logs, sms_logs, payments
# ═════════════════════════════════════════════════════════════════════════

@router.post("/download", summary="Export data to CSV/XLSX")
async def export_download(
    body: ExportRequest,
    session: AsyncSession = Depends(get_db_session),
    tenant_id: UUID = Depends(get_current_tenant_id),
):
    """
    Generate and stream a CSV or XLSX file for the given report type.
    Supports date range filtering, column selection, and simple filters.
    """
    report_type = body.report_type.value

    # Summary reports use a different path
    if report_type in ("funnel_summary", "team_performance", "rfm_breakdown"):
        return await _export_summary(
            session, tenant_id, report_type,
            body.format, body.start_date, body.end_date,
        )

    try:
        col_keys, rows = await fetch_export_rows(
            session=session,
            tenant_id=tenant_id,
            report_type=report_type,
            start_date=body.start_date,
            end_date=body.end_date,
            filters=body.filters,
            columns=body.columns,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if body.format == ExportFormat.csv:
        content = generate_csv_bytes(col_keys, rows, report_type)
        filename = f"{report_type}_{date.today().isoformat()}.csv"
        media_type = "text/csv; charset=utf-8"
    else:
        content = generate_xlsx_bytes(col_keys, rows, report_type, sheet_name=report_type)
        filename = f"{report_type}_{date.today().isoformat()}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        iter([content]),
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Row-Count": str(len(rows)),
        },
    )


async def _export_summary(
    session: AsyncSession,
    tenant_id: UUID,
    report_type: str,
    fmt: ExportFormat,
    start_date: date | None,
    end_date: date | None,
):
    """Handle funnel_summary / team_performance / rfm_breakdown exports."""
    funnel_data = await fetch_funnel_summary(session, tenant_id, start_date, end_date)
    team_data = await fetch_team_performance(session, tenant_id, start_date, end_date)
    rfm_data = await fetch_rfm_breakdown(session, tenant_id)

    if fmt == ExportFormat.pdf:
        content = generate_pdf_bytes(funnel_data, team_data, rfm_data)
        filename = f"report_{date.today().isoformat()}.pdf"
        media_type = "application/pdf"
    elif fmt == ExportFormat.csv:
        # For CSV, export just the requested sub-report
        if report_type == "team_performance" and team_data:
            keys = list(team_data[0].keys())
            content = generate_csv_bytes(keys, team_data, "call_logs")
        elif report_type == "rfm_breakdown" and rfm_data:
            keys = list(rfm_data[0].keys())
            content = generate_csv_bytes(keys, rfm_data, "contacts")
        else:
            keys = ["stage", "count"]
            stage_rows = [
                {"stage": k, "count": v}
                for k, v in (funnel_data.get("latest_stage_counts") or {}).items()
            ]
            content = generate_csv_bytes(keys, stage_rows, "contacts")
        filename = f"{report_type}_{date.today().isoformat()}.csv"
        media_type = "text/csv; charset=utf-8"
    else:
        content = generate_summary_xlsx(funnel_data, team_data, rfm_data)
        filename = f"report_{date.today().isoformat()}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        iter([content]),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ═════════════════════════════════════════════════════════════════════════
# Quick export endpoints (GET-based, for toolbar buttons)
# ═════════════════════════════════════════════════════════════════════════

@router.get("/contacts", summary="Quick export contacts to XLSX")
async def export_contacts(
    session: AsyncSession = Depends(get_db_session),
    tenant_id: UUID = Depends(get_current_tenant_id),
    format: ExportFormat = Query(ExportFormat.xlsx),
    stage: str | None = Query(None),
    segment: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
):
    """Quick-export contacts with optional filters."""
    filters: dict[str, Any] = {}
    if stage:
        filters["current_stage"] = stage
    if segment:
        filters["rfm_segment"] = segment

    col_keys, rows = await fetch_export_rows(
        session, tenant_id, "contacts",
        start_date=start_date, end_date=end_date, filters=filters,
    )
    if format == ExportFormat.csv:
        content = generate_csv_bytes(col_keys, rows, "contacts")
        filename = f"contacts_{date.today().isoformat()}.csv"
        mt = "text/csv; charset=utf-8"
    else:
        content = generate_xlsx_bytes(col_keys, rows, "contacts", "مخاطبین")
        filename = f"contacts_{date.today().isoformat()}.xlsx"
        mt = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        iter([content]),
        media_type=mt,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Row-Count": str(len(rows)),
        },
    )


@router.get("/invoices", summary="Quick export invoices to XLSX")
async def export_invoices(
    session: AsyncSession = Depends(get_db_session),
    tenant_id: UUID = Depends(get_current_tenant_id),
    format: ExportFormat = Query(ExportFormat.xlsx),
    status: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
):
    filters: dict[str, Any] = {}
    if status:
        filters["status"] = status

    col_keys, rows = await fetch_export_rows(
        session, tenant_id, "invoices",
        start_date=start_date, end_date=end_date, filters=filters,
    )
    if format == ExportFormat.csv:
        content = generate_csv_bytes(col_keys, rows, "invoices")
        filename = f"invoices_{date.today().isoformat()}.csv"
        mt = "text/csv; charset=utf-8"
    else:
        content = generate_xlsx_bytes(col_keys, rows, "invoices", "فاکتورها")
        filename = f"invoices_{date.today().isoformat()}.xlsx"
        mt = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        iter([content]),
        media_type=mt,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Row-Count": str(len(rows)),
        },
    )


@router.get("/call-logs", summary="Quick export call logs to XLSX")
async def export_call_logs(
    session: AsyncSession = Depends(get_db_session),
    tenant_id: UUID = Depends(get_current_tenant_id),
    format: ExportFormat = Query(ExportFormat.xlsx),
    status: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
):
    filters: dict[str, Any] = {}
    if status:
        filters["status"] = status

    col_keys, rows = await fetch_export_rows(
        session, tenant_id, "call_logs",
        start_date=start_date, end_date=end_date, filters=filters,
    )
    if format == ExportFormat.csv:
        content = generate_csv_bytes(col_keys, rows, "call_logs")
        filename = f"call_logs_{date.today().isoformat()}.csv"
        mt = "text/csv; charset=utf-8"
    else:
        content = generate_xlsx_bytes(col_keys, rows, "call_logs", "تماس‌ها")
        filename = f"call_logs_{date.today().isoformat()}.xlsx"
        mt = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        iter([content]),
        media_type=mt,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Row-Count": str(len(rows)),
        },
    )


@router.get("/sms-logs", summary="Quick export SMS logs to XLSX")
async def export_sms_logs(
    session: AsyncSession = Depends(get_db_session),
    tenant_id: UUID = Depends(get_current_tenant_id),
    format: ExportFormat = Query(ExportFormat.xlsx),
    status: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
):
    filters: dict[str, Any] = {}
    if status:
        filters["status"] = status

    col_keys, rows = await fetch_export_rows(
        session, tenant_id, "sms_logs",
        start_date=start_date, end_date=end_date, filters=filters,
    )
    if format == ExportFormat.csv:
        content = generate_csv_bytes(col_keys, rows, "sms_logs")
        filename = f"sms_logs_{date.today().isoformat()}.csv"
        mt = "text/csv; charset=utf-8"
    else:
        content = generate_xlsx_bytes(col_keys, rows, "sms_logs", "پیامک‌ها")
        filename = f"sms_logs_{date.today().isoformat()}.xlsx"
        mt = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        iter([content]),
        media_type=mt,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Row-Count": str(len(rows)),
        },
    )


@router.get("/payments", summary="Quick export payments to XLSX")
async def export_payments(
    session: AsyncSession = Depends(get_db_session),
    tenant_id: UUID = Depends(get_current_tenant_id),
    format: ExportFormat = Query(ExportFormat.xlsx),
    status: str | None = Query(None),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
):
    filters: dict[str, Any] = {}
    if status:
        filters["status"] = status

    col_keys, rows = await fetch_export_rows(
        session, tenant_id, "payments",
        start_date=start_date, end_date=end_date, filters=filters,
    )
    if format == ExportFormat.csv:
        content = generate_csv_bytes(col_keys, rows, "payments")
        filename = f"payments_{date.today().isoformat()}.csv"
        mt = "text/csv; charset=utf-8"
    else:
        content = generate_xlsx_bytes(col_keys, rows, "payments", "پرداخت‌ها")
        filename = f"payments_{date.today().isoformat()}.xlsx"
        mt = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        iter([content]),
        media_type=mt,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Row-Count": str(len(rows)),
        },
    )


# ═════════════════════════════════════════════════════════════════════════
# Summary / PDF report
# ═════════════════════════════════════════════════════════════════════════

@router.get("/report/summary", summary="Generate full summary report (PDF/XLSX)")
async def export_summary_report(
    session: AsyncSession = Depends(get_db_session),
    tenant_id: UUID = Depends(get_current_tenant_id),
    format: ExportFormat = Query(ExportFormat.xlsx),
    start_date: date | None = Query(None),
    end_date: date | None = Query(None),
):
    """Generate a comprehensive summary report with funnel, team, and RFM data."""
    return await _export_summary(session, tenant_id, "funnel_summary", format, start_date, end_date)


# ═════════════════════════════════════════════════════════════════════════
# Custom Report Builder
# ═════════════════════════════════════════════════════════════════════════

@router.post("/custom", summary="Build and download a custom report")
async def export_custom_report(
    body: CustomReportRequest,
    session: AsyncSession = Depends(get_db_session),
    tenant_id: UUID = Depends(get_current_tenant_id),
):
    """
    Build a custom report by combining multiple data sources.
    Each data source gets its own sheet in XLSX, or concatenated in CSV.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    if body.format == ExportFormat.pdf:
        raise HTTPException(
            status_code=400,
            detail="PDF format is only available for summary reports. Use XLSX or CSV.",
        )

    all_sheets: list[tuple[str, list[str], list[dict[str, Any]]]] = []

    for src in body.data_sources:
        src_key = src.value
        if src_key in ("funnel_summary", "team_performance", "rfm_breakdown"):
            continue  # Skip summary types in custom builder

        columns_for_src = body.columns.get(src_key) if body.columns else None
        try:
            col_keys, rows = await fetch_export_rows(
                session, tenant_id, src_key,
                start_date=body.start_date, end_date=body.end_date,
                filters=body.filters, columns=columns_for_src,
            )
            all_sheets.append((src_key, col_keys, rows))
        except ValueError:
            continue

    if not all_sheets:
        raise HTTPException(status_code=400, detail="No data sources returned any data")

    if body.format == ExportFormat.csv:
        # Concatenate all sheets into one CSV
        import io, csv as csv_mod
        buf = io.StringIO()
        buf.write("\ufeff")
        writer = csv_mod.writer(buf)
        for sheet_name, col_keys, rows in all_sheets:
            writer.writerow([f"--- {sheet_name} ---"])
            from src.modules.export.service import _get_persian_headers
            headers = _get_persian_headers(sheet_name, col_keys)
            writer.writerow(headers)
            for row in rows:
                writer.writerow([row.get(k, "") for k in col_keys])
            writer.writerow([])
        content = buf.getvalue().encode("utf-8")
        filename = f"custom_report_{date.today().isoformat()}.csv"
        mt = "text/csv; charset=utf-8"
    else:
        # Multi-sheet XLSX
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        header_font = Font(name="B Nazanin", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        data_font = Font(name="B Nazanin", size=10)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        sheet_names_fa = {
            "contacts": "مخاطبین", "invoices": "فاکتورها",
            "call_logs": "تماس‌ها", "sms_logs": "پیامک‌ها",
            "payments": "پرداخت‌ها",
        }

        for sheet_name, col_keys, rows in all_sheets:
            ws = wb.create_sheet(title=sheet_names_fa.get(sheet_name, sheet_name))
            ws.sheet_view.rightToLeft = True

            from src.modules.export.service import _get_persian_headers
            headers = _get_persian_headers(sheet_name, col_keys)

            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border

            for ri, row in enumerate(rows, 2):
                for ci, key in enumerate(col_keys, 1):
                    c = ws.cell(row=ri, column=ci, value=row.get(key, ""))
                    c.font = data_font
                    c.border = thin_border

            for ci in range(1, len(col_keys) + 1):
                letter = get_column_letter(ci)
                max_l = max(
                    len(str(ws.cell(row=r, column=ci).value or ""))
                    for r in range(1, min(len(rows) + 2, 100))
                )
                ws.column_dimensions[letter].width = min(max(max_l + 4, 12), 40)

        import io
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()
        filename = f"custom_report_{date.today().isoformat()}.xlsx"
        mt = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        iter([content]),
        media_type=mt,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


# ═════════════════════════════════════════════════════════════════════════
# Scheduled Reports (in-memory store; replace with DB in production)
# ═════════════════════════════════════════════════════════════════════════

_scheduled_reports: dict[UUID, dict[str, Any]] = {}


@router.get("/schedules", response_model=ScheduledReportListResponse)
async def list_scheduled_reports(
    tenant_id: UUID = Depends(get_current_tenant_id),
):
    """List all scheduled reports for this tenant."""
    items = [
        ScheduledReportResponse(
            id=rid,
            name=r["name"],
            report_type=r["report_type"],
            format=r["format"],
            frequency=r["frequency"],
            recipients=r["recipients"],
            filters=r.get("filters", {}),
            columns=r.get("columns"),
            is_active=r.get("is_active", True),
            last_run_at=r.get("last_run_at"),
            next_run_at=r.get("next_run_at"),
            created_at=r["created_at"],
        )
        for rid, r in _scheduled_reports.items()
        if r.get("tenant_id") == str(tenant_id)
    ]
    return ScheduledReportListResponse(items=items, total=len(items))


@router.post("/schedules", response_model=ScheduledReportResponse, status_code=201)
async def create_scheduled_report(
    body: ScheduledReportRequest,
    tenant_id: UUID = Depends(get_current_tenant_id),
):
    """Create a new scheduled report."""
    rid = uuid4()
    now = datetime.utcnow()

    # Calculate next_run based on frequency
    from datetime import timedelta
    freq_delta = {
        "daily": timedelta(days=1),
        "weekly": timedelta(weeks=1),
        "monthly": timedelta(days=30),
    }
    next_run = now + freq_delta.get(body.frequency.value, timedelta(days=1))

    report = {
        "tenant_id": str(tenant_id),
        "name": body.name,
        "report_type": body.report_type.value,
        "format": body.format.value,
        "frequency": body.frequency.value,
        "recipients": body.recipients,
        "filters": body.filters,
        "columns": body.columns,
        "is_active": body.is_active,
        "last_run_at": None,
        "next_run_at": next_run,
        "created_at": now,
    }
    _scheduled_reports[rid] = report

    return ScheduledReportResponse(
        id=rid,
        name=report["name"],
        report_type=report["report_type"],
        format=report["format"],
        frequency=report["frequency"],
        recipients=report["recipients"],
        filters=report["filters"],
        columns=report["columns"],
        is_active=report["is_active"],
        last_run_at=report["last_run_at"],
        next_run_at=report["next_run_at"],
        created_at=report["created_at"],
    )


@router.delete("/schedules/{schedule_id}", status_code=204)
async def delete_scheduled_report(
    schedule_id: UUID,
    tenant_id: UUID = Depends(get_current_tenant_id),
):
    """Delete a scheduled report."""
    report = _scheduled_reports.get(schedule_id)
    if not report or report.get("tenant_id") != str(tenant_id):
        raise HTTPException(status_code=404, detail="Scheduled report not found")
    del _scheduled_reports[schedule_id]

