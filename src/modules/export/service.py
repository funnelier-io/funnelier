"""
Export & Reporting — Service Layer

Generates CSV, XLSX, and PDF files from database queries.
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import date, datetime, timedelta
from typing import Any, Sequence
from uuid import UUID

import jdatetime
from sqlalchemy import select, func, and_
from sqlalchemy.ext.asyncio import AsyncSession

from src.infrastructure.database.models.leads import ContactModel
from src.infrastructure.database.models.communications import (
    SMSLogModel,
    CallLogModel,
)
from src.infrastructure.database.models.sales import (
    InvoiceModel,
    InvoiceLineItemModel,
    PaymentModel,
)
from src.infrastructure.database.models.analytics import FunnelSnapshotModel

from .schemas import ExportFormat, ReportType

logger = logging.getLogger(__name__)


# ═════════════════════════════════════════════════════════════════════════
# Column definitions per report type (key, English header, Persian header)
# ═════════════════════════════════════════════════════════════════════════

COLUMN_DEFS: dict[str, list[tuple[str, str, str, str]]] = {
    # (key, en_label, fa_label, type)
    "contacts": [
        ("phone_number", "Phone Number", "شماره تلفن", "string"),
        ("name", "Name", "نام", "string"),
        ("email", "Email", "ایمیل", "string"),
        ("source_name", "Source", "منبع", "string"),
        ("category_name", "Category", "دسته‌بندی", "string"),
        ("current_stage", "Funnel Stage", "مرحله فانل", "string"),
        ("rfm_segment", "RFM Segment", "بخش RFM", "string"),
        ("rfm_score", "RFM Score", "امتیاز RFM", "string"),
        ("total_sms_sent", "SMS Sent", "پیامک ارسالی", "number"),
        ("total_sms_delivered", "SMS Delivered", "پیامک تحویلی", "number"),
        ("total_calls", "Total Calls", "کل تماس‌ها", "number"),
        ("total_answered_calls", "Answered Calls", "تماس پاسخ‌داده", "number"),
        ("total_call_duration", "Call Duration (s)", "مدت تماس (ثانیه)", "number"),
        ("total_invoices", "Invoices", "فاکتورها", "number"),
        ("total_paid_invoices", "Paid Invoices", "فاکتور پرداختی", "number"),
        ("total_revenue", "Revenue (Rial)", "درآمد (ریال)", "number"),
        ("tags", "Tags", "برچسب‌ها", "string"),
        ("is_active", "Active", "فعال", "boolean"),
        ("created_at", "Created At", "تاریخ ایجاد", "date"),
    ],
    "invoices": [
        ("invoice_number", "Invoice #", "شماره فاکتور", "string"),
        ("phone_number", "Phone", "شماره تلفن", "string"),
        ("invoice_type", "Type", "نوع", "string"),
        ("status", "Status", "وضعیت", "string"),
        ("subtotal", "Subtotal", "جمع جزئی", "number"),
        ("tax_amount", "Tax", "مالیات", "number"),
        ("discount_amount", "Discount", "تخفیف", "number"),
        ("total_amount", "Total", "مبلغ کل", "number"),
        ("salesperson_name", "Salesperson", "فروشنده", "string"),
        ("issued_at", "Issued At", "تاریخ صدور", "date"),
        ("due_date", "Due Date", "سررسید", "date"),
        ("paid_at", "Paid At", "تاریخ پرداخت", "date"),
        ("created_at", "Created At", "تاریخ ایجاد", "date"),
    ],
    "call_logs": [
        ("phone_number", "Phone", "شماره تلفن", "string"),
        ("call_type", "Type", "نوع", "string"),
        ("source_type", "Source", "منبع", "string"),
        ("salesperson_name", "Salesperson", "فروشنده", "string"),
        ("salesperson_phone", "SP Phone", "شماره فروشنده", "string"),
        ("call_start", "Start", "شروع", "date"),
        ("call_end", "End", "پایان", "date"),
        ("duration_seconds", "Duration (s)", "مدت (ثانیه)", "number"),
        ("status", "Status", "وضعیت", "string"),
        ("is_successful", "Successful", "موفق", "boolean"),
        ("outcome", "Outcome", "نتیجه", "string"),
        ("notes", "Notes", "یادداشت", "string"),
        ("created_at", "Created At", "تاریخ ایجاد", "date"),
    ],
    "sms_logs": [
        ("phone_number", "Phone", "شماره تلفن", "string"),
        ("template_name", "Template", "قالب", "string"),
        ("message_content", "Content", "متن", "string"),
        ("provider", "Provider", "ارائه‌دهنده", "string"),
        ("status", "Status", "وضعیت", "string"),
        ("sent_at", "Sent At", "زمان ارسال", "date"),
        ("delivered_at", "Delivered At", "زمان تحویل", "date"),
        ("cost", "Cost", "هزینه", "number"),
        ("sms_parts", "Parts", "بخش‌ها", "number"),
        ("created_at", "Created At", "تاریخ ایجاد", "date"),
    ],
    "payments": [
        ("phone_number", "Phone", "شماره تلفن", "string"),
        ("amount", "Amount", "مبلغ", "number"),
        ("payment_method", "Method", "روش", "string"),
        ("reference_number", "Reference", "شماره مرجع", "string"),
        ("status", "Status", "وضعیت", "string"),
        ("paid_at", "Paid At", "تاریخ پرداخت", "date"),
        ("created_at", "Created At", "تاریخ ایجاد", "date"),
    ],
}


# ═════════════════════════════════════════════════════════════════════════
# Data fetching
# ═════════════════════════════════════════════════════════════════════════

MODEL_MAP: dict[str, type] = {
    "contacts": ContactModel,
    "invoices": InvoiceModel,
    "call_logs": CallLogModel,
    "sms_logs": SMSLogModel,
    "payments": PaymentModel,
}

# Date column for range filtering
DATE_COLUMN_MAP: dict[str, str] = {
    "contacts": "created_at",
    "invoices": "issued_at",
    "call_logs": "call_start",
    "sms_logs": "sent_at",
    "payments": "paid_at",
}


async def fetch_export_rows(
    session: AsyncSession,
    tenant_id: UUID,
    report_type: str,
    start_date: date | None = None,
    end_date: date | None = None,
    filters: dict[str, Any] | None = None,
    columns: list[str] | None = None,
) -> tuple[list[str], list[dict[str, Any]]]:
    """
    Fetch rows for a given report type, returning (column_keys, rows).
    """
    model = MODEL_MAP.get(report_type)
    if model is None:
        raise ValueError(f"Unsupported report type for tabular export: {report_type}")

    col_defs = COLUMN_DEFS.get(report_type, [])
    if columns:
        col_defs = [c for c in col_defs if c[0] in columns]

    col_keys = [c[0] for c in col_defs]

    # Build query
    stmt = select(model).where(model.tenant_id == tenant_id)  # type: ignore[attr-defined]

    # Date range
    date_col_name = DATE_COLUMN_MAP.get(report_type)
    if date_col_name and start_date:
        date_col = getattr(model, date_col_name, None)
        if date_col is not None:
            stmt = stmt.where(date_col >= datetime.combine(start_date, datetime.min.time()))
    if date_col_name and end_date:
        date_col = getattr(model, date_col_name, None)
        if date_col is not None:
            stmt = stmt.where(date_col <= datetime.combine(end_date, datetime.max.time()))

    # Simple equality filters
    if filters:
        for key, value in filters.items():
            col = getattr(model, key, None)
            if col is not None and value is not None:
                stmt = stmt.where(col == value)

    # Order by created_at desc
    created_col = getattr(model, "created_at", None)
    if created_col is not None:
        stmt = stmt.order_by(created_col.desc())

    result = await session.execute(stmt)
    rows_raw = result.scalars().all()

    rows: list[dict[str, Any]] = []
    for row in rows_raw:
        row_dict: dict[str, Any] = {}
        for key in col_keys:
            val = getattr(row, key, None)
            if isinstance(val, datetime):
                try:
                    jd = jdatetime.datetime.fromgregorian(datetime=val)
                    row_dict[key] = jd.strftime("%Y/%m/%d %H:%M")
                except Exception:
                    row_dict[key] = val.strftime("%Y-%m-%d %H:%M")
            elif isinstance(val, date):
                try:
                    jd = jdatetime.date.fromgregorian(date=val)
                    row_dict[key] = jd.strftime("%Y/%m/%d")
                except Exception:
                    row_dict[key] = val.isoformat()
            elif isinstance(val, (list, dict)):
                row_dict[key] = str(val)
            elif isinstance(val, bool):
                row_dict[key] = "بله" if val else "خیر"
            elif val is None:
                row_dict[key] = ""
            else:
                row_dict[key] = val
        rows.append(row_dict)

    return col_keys, rows


# ═════════════════════════════════════════════════════════════════════════
# Funnel Summary Report data
# ═════════════════════════════════════════════════════════════════════════

async def fetch_funnel_summary(
    session: AsyncSession,
    tenant_id: UUID,
    start_date: date | None = None,
    end_date: date | None = None,
) -> dict[str, Any]:
    """Fetch funnel snapshot data for PDF/Excel summary."""
    from sqlalchemy import desc

    # Get latest snapshot(s) in range
    stmt = select(FunnelSnapshotModel).where(
        FunnelSnapshotModel.tenant_id == tenant_id
    )
    if start_date:
        stmt = stmt.where(FunnelSnapshotModel.snapshot_date >= start_date)
    if end_date:
        stmt = stmt.where(FunnelSnapshotModel.snapshot_date <= end_date)
    stmt = stmt.order_by(desc(FunnelSnapshotModel.snapshot_date)).limit(30)

    result = await session.execute(stmt)
    snapshots = result.scalars().all()

    if not snapshots:
        return {"snapshots": [], "total_revenue": 0, "total_leads": 0}

    latest = snapshots[0]
    total_revenue = sum(s.daily_revenue for s in snapshots)
    total_leads = sum(s.new_leads for s in snapshots)

    return {
        "snapshots": [
            {
                "date": s.snapshot_date.isoformat(),
                "stage_counts": s.stage_counts,
                "new_leads": s.new_leads,
                "daily_revenue": s.daily_revenue,
                "conversion_rates": s.conversion_rates,
            }
            for s in snapshots
        ],
        "latest_stage_counts": latest.stage_counts,
        "latest_conversion_rates": latest.conversion_rates,
        "total_revenue": total_revenue,
        "total_leads": total_leads,
    }


# ═════════════════════════════════════════════════════════════════════════
# Team Performance data
# ═════════════════════════════════════════════════════════════════════════

async def fetch_team_performance(
    session: AsyncSession,
    tenant_id: UUID,
    start_date: date | None = None,
    end_date: date | None = None,
) -> list[dict[str, Any]]:
    """Fetch per-salesperson performance metrics."""
    # Calls per salesperson
    call_stmt = (
        select(
            CallLogModel.salesperson_name,
            func.count(CallLogModel.id).label("total_calls"),
            func.sum(func.cast(CallLogModel.is_successful, Integer)).label("successful_calls"),
            func.sum(CallLogModel.duration_seconds).label("total_duration"),
        )
        .where(CallLogModel.tenant_id == tenant_id)
        .where(CallLogModel.salesperson_name.isnot(None))
    )

    if start_date:
        call_stmt = call_stmt.where(
            CallLogModel.call_start >= datetime.combine(start_date, datetime.min.time())
        )
    if end_date:
        call_stmt = call_stmt.where(
            CallLogModel.call_start <= datetime.combine(end_date, datetime.max.time())
        )

    call_stmt = call_stmt.group_by(CallLogModel.salesperson_name)
    call_result = await session.execute(call_stmt)
    call_rows = call_result.fetchall()

    # Invoices per salesperson
    inv_stmt = (
        select(
            InvoiceModel.salesperson_name,
            func.count(InvoiceModel.id).label("total_invoices"),
            func.sum(InvoiceModel.total_amount).label("total_amount"),
        )
        .where(InvoiceModel.tenant_id == tenant_id)
        .where(InvoiceModel.salesperson_name.isnot(None))
    )
    if start_date:
        inv_stmt = inv_stmt.where(
            InvoiceModel.issued_at >= datetime.combine(start_date, datetime.min.time())
        )
    if end_date:
        inv_stmt = inv_stmt.where(
            InvoiceModel.issued_at <= datetime.combine(end_date, datetime.max.time())
        )
    inv_stmt = inv_stmt.group_by(InvoiceModel.salesperson_name)
    inv_result = await session.execute(inv_stmt)
    inv_rows = inv_result.fetchall()

    # Merge into a dict by salesperson
    perf: dict[str, dict[str, Any]] = {}
    for r in call_rows:
        name = r.salesperson_name or "نامشخص"
        perf.setdefault(name, {"salesperson": name})
        perf[name]["total_calls"] = r.total_calls or 0
        perf[name]["successful_calls"] = r.successful_calls or 0
        perf[name]["total_duration_min"] = round((r.total_duration or 0) / 60, 1)

    for r in inv_rows:
        name = r.salesperson_name or "نامشخص"
        perf.setdefault(name, {"salesperson": name})
        perf[name]["total_invoices"] = r.total_invoices or 0
        perf[name]["total_amount"] = r.total_amount or 0

    return list(perf.values())


# ═════════════════════════════════════════════════════════════════════════
# RFM Breakdown data
# ═════════════════════════════════════════════════════════════════════════

async def fetch_rfm_breakdown(
    session: AsyncSession,
    tenant_id: UUID,
) -> list[dict[str, Any]]:
    """Fetch RFM segment distribution."""
    stmt = (
        select(
            ContactModel.rfm_segment,
            func.count(ContactModel.id).label("count"),
            func.sum(ContactModel.total_revenue).label("revenue"),
            func.avg(ContactModel.total_calls).label("avg_calls"),
        )
        .where(ContactModel.tenant_id == tenant_id)
        .where(ContactModel.rfm_segment.isnot(None))
        .group_by(ContactModel.rfm_segment)
    )
    result = await session.execute(stmt)

    return [
        {
            "segment": r.rfm_segment,
            "count": r.count,
            "revenue": r.revenue or 0,
            "avg_calls": round(float(r.avg_calls or 0), 1),
        }
        for r in result.fetchall()
    ]


# ═════════════════════════════════════════════════════════════════════════
# File generators
# ═════════════════════════════════════════════════════════════════════════

def _get_persian_headers(report_type: str, col_keys: list[str]) -> list[str]:
    """Return Persian headers for given columns."""
    col_defs = COLUMN_DEFS.get(report_type, [])
    lookup = {c[0]: c[2] for c in col_defs}
    return [lookup.get(k, k) for k in col_keys]


def generate_csv_bytes(
    col_keys: list[str],
    rows: list[dict[str, Any]],
    report_type: str,
) -> bytes:
    """Generate a UTF-8 CSV file with BOM for Excel compatibility."""
    buf = io.StringIO()
    buf.write("\ufeff")  # UTF-8 BOM

    headers = _get_persian_headers(report_type, col_keys)
    writer = csv.writer(buf)
    writer.writerow(headers)

    for row in rows:
        writer.writerow([row.get(k, "") for k in col_keys])

    return buf.getvalue().encode("utf-8")


def generate_xlsx_bytes(
    col_keys: list[str],
    rows: list[dict[str, Any]],
    report_type: str,
    sheet_name: str = "Data",
) -> bytes:
    """Generate an XLSX file using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.rightToLeft = True  # RTL for Persian

    headers = _get_persian_headers(report_type, col_keys)

    # Header style
    header_font = Font(name="B Nazanin", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    data_font = Font(name="B Nazanin", size=10)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(col_keys, 1):
            val = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    # Auto-width
    for col_idx in range(1, len(col_keys) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(len(rows) + 2, 100))
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_summary_xlsx(
    funnel_data: dict[str, Any],
    team_data: list[dict[str, Any]],
    rfm_data: list[dict[str, Any]],
    tenant_name: str = "فانلیر",
) -> bytes:
    """Generate a multi-sheet summary XLSX with funnel, team, and RFM tabs."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(name="B Nazanin", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    data_font = Font(name="B Nazanin", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    title_font = Font(name="B Nazanin", bold=True, size=14, color="1E40AF")

    def _style_header(ws, headers, row=1):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border

    def _auto_width(ws, col_count, row_count):
        for ci in range(1, col_count + 1):
            letter = get_column_letter(ci)
            max_l = max(
                len(str(ws.cell(row=r, column=ci).value or ""))
                for r in range(1, min(row_count + 1, 100))
            )
            ws.column_dimensions[letter].width = min(max(max_l + 4, 14), 40)

    # ── Sheet 1: Funnel Summary ──
    ws1 = wb.active
    ws1.title = "خلاصه فانل"
    ws1.sheet_view.rightToLeft = True
    ws1.cell(row=1, column=1, value=f"گزارش فانل فروش — {tenant_name}").font = title_font
    ws1.merge_cells("A1:G1")

    stage_counts = funnel_data.get("latest_stage_counts", {})
    stage_labels = {
        "lead_acquired": "سرنخ",
        "sms_sent": "پیامک ارسالی",
        "sms_delivered": "تحویل پیامک",
        "call_attempted": "تماس",
        "call_answered": "پاسخ تماس",
        "invoice_issued": "پیش‌فاکتور",
        "payment_received": "پرداخت",
    }
    _style_header(ws1, ["مرحله", "تعداد"], row=3)
    for idx, (key, label) in enumerate(stage_labels.items()):
        r = idx + 4
        ws1.cell(row=r, column=1, value=label).font = data_font
        ws1.cell(row=r, column=2, value=stage_counts.get(key, 0)).font = data_font
        ws1.cell(row=r, column=1).border = thin_border
        ws1.cell(row=r, column=2).border = thin_border

    total_row = len(stage_labels) + 5
    ws1.cell(row=total_row, column=1, value="کل درآمد (ریال)").font = Font(name="B Nazanin", bold=True, size=11)
    ws1.cell(row=total_row, column=2, value=funnel_data.get("total_revenue", 0)).font = data_font
    _auto_width(ws1, 2, total_row)

    # ── Sheet 2: Team Performance ──
    ws2 = wb.create_sheet(title="عملکرد تیم")
    ws2.sheet_view.rightToLeft = True
    team_headers = ["فروشنده", "تماس‌ها", "تماس موفق", "مدت (دقیقه)", "فاکتورها", "مبلغ فروش (ریال)"]
    _style_header(ws2, team_headers)
    for idx, t in enumerate(team_data, 2):
        ws2.cell(row=idx, column=1, value=t.get("salesperson", "")).font = data_font
        ws2.cell(row=idx, column=2, value=t.get("total_calls", 0)).font = data_font
        ws2.cell(row=idx, column=3, value=t.get("successful_calls", 0)).font = data_font
        ws2.cell(row=idx, column=4, value=t.get("total_duration_min", 0)).font = data_font
        ws2.cell(row=idx, column=5, value=t.get("total_invoices", 0)).font = data_font
        ws2.cell(row=idx, column=6, value=t.get("total_amount", 0)).font = data_font
        for ci in range(1, 7):
            ws2.cell(row=idx, column=ci).border = thin_border
    _auto_width(ws2, 6, len(team_data) + 1)

    # ── Sheet 3: RFM Breakdown ──
    ws3 = wb.create_sheet(title="بخش‌بندی RFM")
    ws3.sheet_view.rightToLeft = True
    rfm_headers = ["بخش", "تعداد", "درآمد (ریال)", "میانگین تماس"]
    segment_labels = {
        "champions": "قهرمانان", "loyal": "وفادار",
        "potential_loyalist": "بالقوه وفادار", "new_customers": "مشتریان جدید",
        "promising": "امیدوار", "need_attention": "نیاز به توجه",
        "about_to_sleep": "رو به خواب", "at_risk": "در خطر",
        "cant_lose": "از دست ندهید", "hibernating": "خواب",
        "lost": "از دست رفته",
    }
    _style_header(ws3, rfm_headers)
    for idx, r in enumerate(rfm_data, 2):
        seg = r.get("segment", "")
        ws3.cell(row=idx, column=1, value=segment_labels.get(seg, seg)).font = data_font
        ws3.cell(row=idx, column=2, value=r.get("count", 0)).font = data_font
        ws3.cell(row=idx, column=3, value=r.get("revenue", 0)).font = data_font
        ws3.cell(row=idx, column=4, value=r.get("avg_calls", 0)).font = data_font
        for ci in range(1, 5):
            ws3.cell(row=idx, column=ci).border = thin_border
    _auto_width(ws3, 4, len(rfm_data) + 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_pdf_bytes(
    funnel_data: dict[str, Any],
    team_data: list[dict[str, Any]],
    rfm_data: list[dict[str, Any]],
    tenant_name: str = "فانلیر",
    report_date: str | None = None,
) -> bytes:
    """
    Generate a simple PDF report.
    Uses a simple HTML-to-PDF approach via a minimal table layout.
    Falls back to CSV if reportlab is not available.
    """
    if report_date is None:
        try:
            jd = jdatetime.date.today()
            report_date = jd.strftime("%Y/%m/%d")
        except Exception:
            report_date = date.today().isoformat()

    try:
        return _generate_pdf_reportlab(funnel_data, team_data, rfm_data, tenant_name, report_date)
    except ImportError:
        logger.warning("reportlab not installed — falling back to text-based PDF")
        return _generate_pdf_text_fallback(funnel_data, team_data, rfm_data, tenant_name, report_date)


def _generate_pdf_reportlab(
    funnel_data: dict[str, Any],
    team_data: list[dict[str, Any]],
    rfm_data: list[dict[str, Any]],
    tenant_name: str,
    report_date: str,
) -> bytes:
    """Generate PDF using reportlab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(
        f"Funnelier Report — {tenant_name} — {report_date}",
        styles["Title"],
    ))
    elements.append(Spacer(1, 10 * mm))

    # Funnel summary table
    stage_labels = {
        "lead_acquired": "Lead Acquired",
        "sms_sent": "SMS Sent",
        "sms_delivered": "SMS Delivered",
        "call_attempted": "Call Attempted",
        "call_answered": "Call Answered",
        "invoice_issued": "Invoice Issued",
        "payment_received": "Payment Received",
    }
    stage_counts = funnel_data.get("latest_stage_counts", {})
    funnel_table_data = [["Stage", "Count"]]
    for key, label in stage_labels.items():
        funnel_table_data.append([label, str(stage_counts.get(key, 0))])
    funnel_table_data.append(["Total Revenue (Rial)", str(funnel_data.get("total_revenue", 0))])

    t = Table(funnel_table_data, colWidths=[120 * mm, 50 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
    ]))
    elements.append(Paragraph("Funnel Summary", styles["Heading2"]))
    elements.append(t)
    elements.append(Spacer(1, 8 * mm))

    # Team performance table
    if team_data:
        team_table_data = [["Salesperson", "Calls", "Successful", "Duration (min)", "Invoices", "Amount"]]
        for tp in team_data:
            team_table_data.append([
                tp.get("salesperson", ""),
                str(tp.get("total_calls", 0)),
                str(tp.get("successful_calls", 0)),
                str(tp.get("total_duration_min", 0)),
                str(tp.get("total_invoices", 0)),
                str(tp.get("total_amount", 0)),
            ])
        t2 = Table(team_table_data)
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#059669")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ]))
        elements.append(Paragraph("Team Performance", styles["Heading2"]))
        elements.append(t2)
        elements.append(Spacer(1, 8 * mm))

    # RFM breakdown table
    if rfm_data:
        rfm_table_data = [["Segment", "Count", "Revenue", "Avg Calls"]]
        for r in rfm_data:
            rfm_table_data.append([
                r.get("segment", ""),
                str(r.get("count", 0)),
                str(r.get("revenue", 0)),
                str(r.get("avg_calls", 0)),
            ])
        t3 = Table(rfm_table_data)
        t3.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#7C3AED")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ]))
        elements.append(Paragraph("RFM Breakdown", styles["Heading2"]))
        elements.append(t3)

    doc.build(elements)
    return buf.getvalue()


def _generate_pdf_text_fallback(
    funnel_data: dict[str, Any],
    team_data: list[dict[str, Any]],
    rfm_data: list[dict[str, Any]],
    tenant_name: str,
    report_date: str,
) -> bytes:
    """Simple text fallback when reportlab is unavailable."""
    lines = [
        f"Funnelier Report — {tenant_name}",
        f"Date: {report_date}",
        "=" * 50,
        "",
        "FUNNEL SUMMARY",
        "-" * 30,
    ]
    for stage, count in (funnel_data.get("latest_stage_counts") or {}).items():
        lines.append(f"  {stage}: {count}")
    lines.append(f"  Total Revenue: {funnel_data.get('total_revenue', 0)}")
    lines.append("")

    if team_data:
        lines.append("TEAM PERFORMANCE")
        lines.append("-" * 30)
        for tp in team_data:
            lines.append(
                f"  {tp.get('salesperson', 'N/A')}: "
                f"calls={tp.get('total_calls', 0)}, "
                f"invoices={tp.get('total_invoices', 0)}, "
                f"amount={tp.get('total_amount', 0)}"
            )
        lines.append("")

    if rfm_data:
        lines.append("RFM BREAKDOWN")
        lines.append("-" * 30)
        for r in rfm_data:
            lines.append(
                f"  {r.get('segment', 'N/A')}: "
                f"count={r.get('count', 0)}, "
                f"revenue={r.get('revenue', 0)}"
            )

    return "\n".join(lines).encode("utf-8")


# ═════════════════════════════════════════════════════════════════════════
# Integer helpers from sqlalchemy
# ═════════════════════════════════════════════════════════════════════════

from sqlalchemy import Integer

